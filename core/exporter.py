import csv
import json
from datetime import datetime
from pathlib import Path

from core.database import db


class DataExporter:
    """数据导出器"""
    
    @staticmethod
    def export_to_csv(documents, output_path, include_content=True):
        """导出文档到CSV
        
        Args:
            documents: 文档列表
            output_path: 输出文件路径
            include_content: 是否包含正文内容
            
        Returns:
            dict: 导出结果
        """
        try:
            fieldnames = ['id', 'title', 'file_type', 'file_size', 'year', 
                         'author', 'source', 'category', 'tags', 'import_date']
            
            if include_content:
                fieldnames.append('content_text')
            
            # 获取分类映射
            categories = {cat['id']: cat['name'] for cat in db.get_categories()}
            
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                
                for doc in documents:
                    row = {
                        'id': doc['id'],
                        'title': doc['title'] or '',
                        'file_type': doc['file_type'] or '',
                        'file_size': doc['file_size'] or 0,
                        'year': doc['year'] or '',
                        'author': doc['author'] or '',
                        'source': doc['source'] or '',
                        'category': categories.get(doc['category_id'], '未分类'),
                        'tags': ', '.join([t['name'] for t in db.get_document_tags(doc['id'])]),
                        'import_date': doc['import_date'] or ''
                    }
                    
                    if include_content:
                        row['content_text'] = doc['content_text'] or ''
                    
                    writer.writerow(row)
            
            return {
                'success': True,
                'count': len(documents),
                'path': output_path
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def export_to_excel(documents, output_path, include_content=False):
        """导出文档到Excel
        
        Args:
            documents: 文档列表
            output_path: 输出文件路径
            include_content: 是否包含正文内容
            
        Returns:
            dict: 导出结果
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "文档列表"
            
            # 设置表头
            headers = ['ID', '标题', '类型', '大小(字节)', '年份', 
                      '作者', '来源', '分类', '标签', '导入时间']
            
            if include_content:
                headers.append('正文内容')
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # 获取分类映射
            categories = {cat['id']: cat['name'] for cat in db.get_categories()}
            
            # 写入数据
            for row_idx, doc in enumerate(documents, 2):
                ws.cell(row=row_idx, column=1, value=doc['id'])
                ws.cell(row=row_idx, column=2, value=doc['title'] or '')
                ws.cell(row=row_idx, column=3, value=doc['file_type'] or '')
                ws.cell(row=row_idx, column=4, value=doc['file_size'] or 0)
                ws.cell(row=row_idx, column=5, value=doc['year'] or '')
                ws.cell(row=row_idx, column=6, value=doc['author'] or '')
                ws.cell(row=row_idx, column=7, value=doc['source'] or '')
                ws.cell(row=row_idx, column=8, value=categories.get(doc['category_id'], '未分类'))
                ws.cell(row=row_idx, column=9, value=', '.join([t['name'] for t in db.get_document_tags(doc['id'])]))
                ws.cell(row=row_idx, column=10, value=str(doc['import_date']) if doc['import_date'] else '')
                
                if include_content:
                    ws.cell(row=row_idx, column=11, value=doc['content_text'] or '')
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            
            if include_content:
                ws.column_dimensions['K'].width = 50
            
            wb.save(output_path)
            
            return {
                'success': True,
                'count': len(documents),
                'path': output_path
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def export_quotes_to_excel(quotes, output_path):
        """导出金句到Excel
        
        Args:
            quotes: 金句列表
            output_path: 输出文件路径
            
        Returns:
            dict: 导出结果
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "金句库"
            
            # 设置表头
            headers = ['ID', '内容', '分类', '标签', '创建时间', '备注']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # 写入数据
            for row_idx, quote in enumerate(quotes, 2):
                ws.cell(row=row_idx, column=1, value=quote['id'])
                ws.cell(row=row_idx, column=2, value=quote['content'] or '')
                ws.cell(row=row_idx, column=3, value=quote['category'] or '')
                ws.cell(row=row_idx, column=4, value=quote['tags'] or '')
                ws.cell(row=row_idx, column=5, value=str(quote['created_date']) if quote['created_date'] else '')
                ws.cell(row=row_idx, column=6, value=quote['notes'] or '')
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 30
            
            wb.save(output_path)
            
            return {
                'success': True,
                'count': len(quotes),
                'path': output_path
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    @staticmethod
    def export_statistics(output_path):
        """导出系统统计信息到JSON
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            dict: 导出结果
        """
        try:
            # 收集统计信息
            documents = db.get_all_documents()
            categories = db.get_categories()
            tags = db.get_tags()
            quotes = db.get_quotes()
            
            # 按年份统计
            year_stats = {}
            for doc in documents:
                year = doc.get('year') or '未知'
                year_stats[year] = year_stats.get(year, 0) + 1
            
            # 按类型统计
            type_stats = {}
            for doc in documents:
                file_type = doc.get('file_type') or '未知'
                type_stats[file_type] = type_stats.get(file_type, 0) + 1
            
            # 按分类统计
            category_stats = {}
            cat_map = {cat['id']: cat['name'] for cat in categories}
            for doc in documents:
                cat_id = doc.get('category_id')
                cat_name = cat_map.get(cat_id, '未分类')
                category_stats[cat_name] = category_stats.get(cat_name, 0) + 1
            
            stats = {
                'export_time': datetime.now().isoformat(),
                'summary': {
                    'total_documents': len(documents),
                    'total_categories': len(categories),
                    'total_tags': len(tags),
                    'total_quotes': len(quotes)
                },
                'year_distribution': year_stats,
                'type_distribution': type_stats,
                'category_distribution': category_stats
            }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(stats, f, ensure_ascii=False, indent=2)
            
            return {
                'success': True,
                'path': output_path
            }
            
        except Exception as e:
            return {'success': False, 'error': str(e)}


# 便捷函数
def export_documents_csv(documents, output_path, include_content=True):
    """导出文档到CSV的便捷函数"""
    return DataExporter.export_to_csv(documents, output_path, include_content)


def export_documents_excel(documents, output_path, include_content=False):
    """导出文档到Excel的便捷函数"""
    return DataExporter.export_to_excel(documents, output_path, include_content)
